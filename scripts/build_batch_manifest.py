#!/usr/bin/env python3
"""scripts/build_batch_manifest.py

Walk every Batch_*/ directory under inputs/TSO500-HRD/, parse the curated
"定序檢體清單"/"重點摘要" xlsx in each, and emit a single master manifest TSV
with columns: sample_id, batch, tumor_zh, oncokb_code, tsv_path.

The OncoKB code mapping is inline below; extend as new tumor names appear.
"""
import os, re, sys, glob
from pathlib import Path
import openpyxl

ROOT = Path("inputs/TSO500-HRD")
MANIFEST = Path("inputs/TSO500-HRD/_manifest.tsv")

# Chinese cancer type → OncoKB code mapping
ZH_TO_ONCOKB = {
    "子宮內膜癌": "UCEC", "endometrial ca": "UCEC", "endometrial cancer": "UCEC",
    "子宮體癌": "UCEC",  # uterine corpus cancer = endometrial
    "乳癌": "BRCA", "乳癌左乳": "BRCA", "乳癌右乳": "BRCA",
    "大腸癌": "COAD", "結腸癌": "COAD", "乙狀結腸": "COAD",
    "直腸癌": "READ", "直腸": "READ",
    "肺癌": "NSCLC", "肺腺癌": "LUAD", "非小細胞肺癌": "NSCLC", "小細胞肺癌": "SCLC",
    "卵巢癌": "OV", "高分化漿液性卵巢癌": "HGSOC", "漿液性卵巢癌": "HGSOC",
    "胰臟癌": "PAAD", "胰臟": "PAAD",
    "胃癌": "STAD", "胃腺癌": "STAD",
    "肝癌": "LIHC", "肝細胞癌": "LIHC", "肝細胞": "LIHC",
    "前列腺癌": "PRAD", "攝護腺癌": "PRAD",
    "食道癌": "ESCA",
    "膽囊癌": "CHOL", "膽管癌": "CHOL", "肝外膽管癌": "CHOL",
    "黑色素瘤": "MEL", "皮膚癌": "MEL",
    "腎癌": "RCC", "腎細胞癌": "RCC",
    "子宮頸癌": "CESC", "頸癌": "CESC",
    "甲狀腺癌": "THYC",
    "肉瘤": "SARC", "軟組織肉瘤": "SARC", "sarcoma": "SARC",
    "cervical ca": "CESC", "cervical cancer": "CESC",
    "口腔癌": "HNSC", "頭頸癌": "HNSC",
    "膀胱癌": "BLCA",
    "腦癌": "GBM", "膠質母細胞瘤": "GBM",
    "白血病": "AML",
    "淋巴癌": "DLBCL",
    "腹膜癌": "ICAR",
}

SAMPLE_RE = re.compile(r"^M\d{2}-\d{4}R$")

def normalise(s):
    if s is None: return ""
    return str(s).strip().replace("　", " ")

def map_tumor(zh):
    """Map a Chinese cancer description to OncoKB code; case-insensitive partial match."""
    if not zh: return ""
    zh_l = zh.strip().lower()
    # Drop common modifiers that don't affect cancer-type mapping
    zh_l = re.sub(r"\s*(left|right|metastatic|recurrent|advanced|primary|tumor)\s*",
                  " ", zh_l).strip()
    # Direct hit (exact, case-insensitive)
    for k, v in ZH_TO_ONCOKB.items():
        if k.lower() == zh_l:
            return v
    # Partial — longest key wins (case-insensitive)
    matches = [(k, v) for k, v in ZH_TO_ONCOKB.items()
               if k in zh or k.lower() in zh_l]
    if matches:
        return max(matches, key=lambda x: len(x[0]))[1]
    return ""

CANCER_HEADER_EXACT = ("癌別", "癌種")          # Chinese — exact match
CANCER_HEADER_WORDS = ("Cancer", "Diagnosis")    # English — whole-word match
NOTES_HEADER_HINTS = ("Notes", "備註", "Note")

def parse_xlsx(path):
    """Return list of (sample_id, tumor_zh) tuples from a 定序檢體清單 / 重點摘要 sheet.

    Tries (a) a column named 癌別/癌種/Cancer/etc., (b) falls back to the
    Notes column when those are missing — Batch 2 inlines the cancer type
    in Notes.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    seen = set()
    for ws in wb.worksheets:
        for header_row in range(1, min(8, ws.max_row + 1)):
            headers = [normalise(c.value) for c in ws[header_row]]
            if any("Sample ID" in h for h in headers):
                sid_col = next(i for i, h in enumerate(headers) if "Sample ID" in h)
                cancer_col = next((i for i, h in enumerate(headers)
                                   if any(hint in h for hint in CANCER_HEADER_EXACT)
                                   or any(re.search(rf"\b{hint}\b", h, re.I)
                                          for hint in CANCER_HEADER_WORDS)), None)
                notes_col = next((i for i, h in enumerate(headers)
                                  if any(hint in h or hint.lower() in h.lower()
                                         for hint in NOTES_HEADER_HINTS)), None)
                if cancer_col is None and notes_col is None:
                    break
                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    sid = normalise(row[sid_col]) if sid_col < len(row) else ""
                    if not SAMPLE_RE.match(sid): continue
                    zh = ""
                    if cancer_col is not None and cancer_col < len(row):
                        zh = normalise(row[cancer_col])
                    if not zh and notes_col is not None and notes_col < len(row):
                        zh = normalise(row[notes_col])
                    if sid in seen: continue
                    seen.add(sid)
                    rows.append((sid, zh))
                break
    return rows

def main():
    out_lines = ["sample_id\tbatch\ttumor_zh\toncokb_code\ttsv_path"]
    n_per_batch = {}
    for batch_dir in sorted(ROOT.glob("Batch_*")):
        batch = batch_dir.name
        # Try every xlsx in the batch dir; combine results
        seen = {}
        for xlsx in batch_dir.glob("*.xlsx"):
            if "MetricsOutput" in xlsx.name: continue
            try:
                for sid, zh in parse_xlsx(xlsx):
                    if sid not in seen:
                        seen[sid] = zh
            except Exception as e:
                print(f"  WARN parse {xlsx}: {e}", file=sys.stderr)
        # Cross-reference against actual TSVs in the directory
        for tsv in sorted(batch_dir.glob("*_CombinedVariantOutput.tsv")):
            sid = tsv.name.split("_CombinedVariantOutput")[0]
            zh = seen.get(sid, "")
            code = map_tumor(zh)
            out_lines.append(f"{sid}\t{batch}\t{zh}\t{code}\t{tsv}")
        n_per_batch[batch] = len(list(batch_dir.glob("*_CombinedVariantOutput.tsv")))

    MANIFEST.write_text("\n".join(out_lines) + "\n")
    print(f"Wrote {MANIFEST} with {len(out_lines)-1} rows.")
    for b, n in sorted(n_per_batch.items()):
        print(f"  {b}: {n} samples")

if __name__ == "__main__":
    main()
